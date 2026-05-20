from typing import Optional, Dict, Literal

import numpy as np
import pandas as pd
import re
from openpyxl import load_workbook
from pathlib import Path

from qresearch.data.catalog import load_dataset_metadata, resolve_dataset_path
from qresearch.data.jqdata import build_jq_panel_parquet
from qresearch.data.types import MarketData
from qresearch.data.utils import get_raw_dir, get_processed_dir


SECTOR_RE = re.compile(r"恒生綜合行業指數\s*-\s*(.*?)\s*之成份股變動")

GROUP_COL_CANDIDATES = [
    "Effective Date 生效日期",
    "No. of Constituents",
    "Change 變動",
    "Count 數目",
]

ID_COL = "Stock Code 股份代號"  # used to identify real data rows


def marketdata_to_panel(md: MarketData) -> pd.DataFrame:
    """
    Convert MarketData into the canonical wide panel format.

    Layout:
      - index: DatetimeIndex
      - columns: MultiIndex (Field, Ticker)
    """
    parts: dict[str, pd.DataFrame] = {"Close": md.close}
    if md.open is not None:
        parts["Open"] = md.open
    if md.high is not None:
        parts["High"] = md.high
    if md.low is not None:
        parts["Low"] = md.low
    if md.volume is not None:
        parts["Volume"] = md.volume
    if md.turnover is not None:
        parts["Turnover"] = md.turnover

    panel = pd.concat(parts, axis=1).sort_index(axis=1)
    if not isinstance(panel.index, pd.DatetimeIndex):
        panel.index = pd.to_datetime(panel.index)
    panel.index = panel.index.sort_values()
    panel = panel.sort_index()
    return panel


def panel_to_marketdata(panel: pd.DataFrame) -> MarketData:
    """
    Convert a canonical wide panel into MarketData.
    """
    if not isinstance(panel.index, pd.DatetimeIndex):
        panel = panel.copy()
        panel.index = pd.to_datetime(panel.index)
    if not isinstance(panel.columns, pd.MultiIndex) or panel.columns.nlevels != 2:
        raise ValueError("Panel must contain MultiIndex columns (Field, Ticker).")

    panel = panel.sort_index().sort_index(axis=1)
    inferred_freq = pd.infer_freq(panel.index) if len(panel.index) >= 3 else None
    if inferred_freq is not None:
        panel.index.freq = inferred_freq

    lvl0 = panel.columns.get_level_values(0)

    def _get(field: str) -> pd.DataFrame | None:
        return panel[field].copy() if field in lvl0 else None

    return MarketData(
        close=_get("Close"),
        open=_get("Open"),
        high=_get("High"),
        low=_get("Low"),
        volume=_get("Volume"),
        turnover=_get("Turnover"),
    )


def save_market_data_to_parquet(md: MarketData, file_path: Path | str) -> None:
    """
    Save MarketData as a single Parquet file in yfinance-style MultiIndex columns:
      level0 = Field (Close/Open/High/Low/Volume)
      level1 = Ticker
    Includes only fields that exist (not None).
    """
    panel = marketdata_to_panel(md)

    file_path = Path(file_path)
    file_path.parent.mkdir(parents=True, exist_ok=True)

    panel.to_parquet(file_path)  # requires pyarrow or fastparquet installed


def _resolve_artifact_path(file_path: Path | str, *, root: Path | str | None = None) -> Path:
    path = Path(file_path)
    if path.exists():
        return path
    if path.suffix:
        return path
    return resolve_dataset_path(str(file_path), root=root)


def load_market_data_from_parquet(file_path: Path | str, *, root: Path | str | None = None) -> MarketData:
    """
    Load a **canonical MarketData panel** from a Parquet file.

    Expected Parquet layout (wide "panel" format):
      - index: DatetimeIndex of trading dates (daily frequency)
      - columns: MultiIndex with exactly 2 levels:
          level 0 = Field name (e.g., "Open", "Close", "High", "Low", "Volume", "Turnover")
          level 1 = Ticker / instrument identifier (e.g., "600519.XSHG")

    This format is intentionally compatible with a "yfinance-style" panel, but is used as
    the internal canonical storage for *any* data source (yfinance / JoinQuant / etc.).

    Behavior:
      - If the index is not a DatetimeIndex, it is coerced via pd.to_datetime.
      - Missing fields are returned as None in the MarketData object.
      - No alignment, forward-filling, or corporate-action adjustments are applied here.
        Those should be handled upstream (data processing) or downstream (research logic).
    """
    file_path = _resolve_artifact_path(file_path, root=root)
    panel = pd.read_parquet(file_path)

    return panel_to_marketdata(panel)


def save_market_data_to_csv(md: MarketData, file_path: Path | str) -> None:
    """
    Save MarketData to CSV using the same canonical panel format as parquet storage.
    """
    panel = marketdata_to_panel(md)
    file_path = Path(file_path)
    file_path.parent.mkdir(parents=True, exist_ok=True)
    panel.to_csv(file_path)


def load_market_data_from_csv(file_path: Path | str, *, root: Path | str | None = None) -> MarketData:
    """
    Load MarketData from a CSV saved in the canonical panel format.
    """
    resolved = _resolve_artifact_path(file_path, root=root)
    panel = pd.read_csv(resolved, header=[0, 1], index_col=0, parse_dates=True)
    return panel_to_marketdata(panel)


def load_market_data_with_metadata(
    file_path: Path | str,
    *,
    root: Path | str | None = None,
) -> tuple[MarketData, object | None]:
    resolved = _resolve_artifact_path(file_path, root=root)
    suffix = resolved.suffix.lower()
    if suffix == ".parquet":
        md = load_market_data_from_parquet(resolved, root=root)
    elif suffix == ".csv":
        md = load_market_data_from_csv(resolved, root=root)
    else:
        raise ValueError(f"Unsupported market-data artifact type: {resolved.suffix}")
    metadata = load_dataset_metadata(resolved, root=root)
    return md, metadata


def load_limit_daily_with_open(
    *,
    file_path: Path | str,
    dates: pd.DatetimeIndex,
    tickers: pd.Index,
) -> Dict[str, pd.DataFrame]:
    """
    Load processed limit_daily parquet (already contains 'open') and pivot into matrices:
      open, close, volume, high_limit, low_limit

    Returns matrices aligned to (dates x tickers).
    """
    file_path = Path(file_path)
    df = pd.read_parquet(file_path)

    # normalize keys
    df = df.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.normalize()
    df["code"] = df["code"].astype("string").str.strip()

    # keep only needed cols (safe if extras exist)
    keep = ["date", "code", "open", "close", "volume", "high_limit", "low_limit"]
    missing = [c for c in keep if c not in df.columns]
    if missing:
        raise ValueError(f"{file_path.name} missing columns: {missing}. Got: {list(df.columns)}")

    out = {
        "open": long_to_matrix(df, date_col="date", code_col="code", value_col="open", dates=dates, tickers=tickers),
        "close": long_to_matrix(df, date_col="date", code_col="code", value_col="close", dates=dates, tickers=tickers),
        "volume": long_to_matrix(df, date_col="date", code_col="code", value_col="volume", dates=dates, tickers=tickers),
        "high_limit": long_to_matrix(df, date_col="date", code_col="code", value_col="high_limit", dates=dates, tickers=tickers),
        "low_limit": long_to_matrix(df, date_col="date", code_col="code", value_col="low_limit", dates=dates, tickers=tickers),
    }
    return out


def load_jq_market_data(
    jq_prices_parquet: Path | str,
    panel_cache_path: Path | str,
) -> "MarketData":
    panel_cache_path = Path(panel_cache_path)

    if not panel_cache_path.exists():
        build_jq_panel_parquet(
            jq_prices_parquet=jq_prices_parquet,
            out_panel_parquet=panel_cache_path,
        )

    return load_market_data_from_parquet(panel_cache_path)


# 你可以在这里扩展/改别名：用“人类可读名称”来选用哪套行业口径
SECTOR_MAP_ALIASES = {
    # 申万
    "sw一级": "sw_l1",
    "sw1": "sw_l1",
    "申万一级": "sw_l1",
    "sw二级": "sw_l2",
    "sw2": "sw_l2",
    "申万二级": "sw_l2",
    "sw三级": "sw_l3",
    "sw3": "sw_l3",
    "申万三级": "sw_l3",

    # 聚宽
    "jq一级": "jq_l1",
    "聚宽一级": "jq_l1",
    "jq二级": "jq_l2",
    "聚宽二级": "jq_l2",

    # 证监会
    "证监会": "zjw",
    "zjw": "zjw",
}

def load_sector_mapper(
    csv_path: str = get_processed_dir() / 'jqdata' / 'industry_map.csv',
    taxonomy: str = '申万一级',
    *,
    code_col: str = "stock_code",
    prefer: Literal["name", "code"] = "name",
    return_series: bool = True,
    dropna: bool = True,
) -> "Dict[str, str] | pd.Series":
    """
    从本地行业映射 CSV 里，构建一个 {stock_code -> sector_name} 的映射（或返回 Series）。

    约定：你的 CSV 是“flatten 过”的形式，至少包含：
      - stock_code
      - <tax>_name / <tax>_code  （例如 sw_l1_name, sw_l1_code）

    参数
    - taxonomy: 用名称选择行业口径（例如 "sw一级", "申万二级", "jq一级", "证监会"）
    - prefer: 默认返回 name（行业中文名）；也可以选 "code"
    - return_series: True 则返回 pd.Series（index=stock_code, value=sector）
    - dropna: True 则丢掉 sector 为空的行

    返回
    - dict: { '000001.XSHE': '银行', ... }  或  Series
    """
    tax_key = SECTOR_MAP_ALIASES.get(taxonomy.strip().lower(), None)
    if tax_key is None:
        # 也允许你直接传 sw_l1/jq_l2 这种“内部key”
        tax_key = taxonomy.strip()

    df = pd.read_csv(csv_path)

    # 自动识别列名（兼容你可能还没改名的情况）
    # 期望：sw_l1_name / sw_l1_code 这种
    want_col = f"{tax_key}_{prefer}"
    if want_col not in df.columns:
        # 兼容旧命名：sw_l1_industry_name / sw_l1_industry_code
        alt = f"{tax_key}_industry_{prefer}"
        if alt in df.columns:
            want_col = alt
        else:
            raise KeyError(
                f"找不到行业列：期望 {tax_key}_{prefer} 或 {tax_key}_industry_{prefer}，"
                f"但实际列为：{list(df.columns)}"
            )

    if code_col not in df.columns:
        raise KeyError(f"找不到股票代码列 {code_col}，实际列为：{list(df.columns)}")

    out = df[[code_col, want_col]].copy()
    out = out.drop_duplicates(subset=[code_col], keep="last")
    if dropna:
        out = out[out[want_col].notna() & (out[want_col].astype(str).str.len() > 0)]

    s = pd.Series(out[want_col].values, index=out[code_col].values, name=want_col)

    return s if return_series else s.to_dict()


def save_historical_mcap_parquet(
    mcap: pd.DataFrame,
    file_path: str | Path,
    *,
    sort_index: bool = True,
    sort_columns: bool = True,
    coerce_float32: bool = False,
    compression: str = "zstd",
) -> Path:
    """
    Save historical market cap panel (date x ticker) to Parquet.

    Assumptions:
      - index is dates (or convertible to DatetimeIndex)
      - columns are tickers (strings)
      - values are numeric (floats), NaNs allowed

    Args:
      mcap: DataFrame indexed by date with ticker columns.
      file_path: destination .parquet path.
      sort_index: sort dates ascending before saving.
      sort_columns: sort tickers lexicographically before saving.
      coerce_float32: optionally downcast to float32 to reduce size (may lose precision).
      compression: parquet compression ("zstd", "snappy", "gzip", etc.).

    Returns:
      Path to saved file.
    """
    fp = Path(file_path)
    fp.parent.mkdir(parents=True, exist_ok=True)

    df = mcap.copy()

    # Ensure datetime index
    if not isinstance(df.index, pd.DatetimeIndex):
        df.index = pd.to_datetime(df.index, errors="raise")

    # Optional sorting for stable diffs / reproducibility
    if sort_index:
        df = df.sort_index()
    if sort_columns:
        df = df.reindex(sorted(df.columns), axis=1)

    # Ensure numeric (keep NaNs)
    df = df.apply(pd.to_numeric, errors="coerce")

    # Optional downcast
    if coerce_float32:
        df = df.astype("float32")

    # Save
    df.to_parquet(fp, compression=compression)

    return fp

def load_historical_mcap_parquet(
    file_path: str | Path,
    *,
    tickers: list[str] | None = None,
    start: str | pd.Timestamp | None = None,
    end: str | pd.Timestamp | None = None,
    sort_index: bool = True,
    sort_columns: bool = False,
) -> pd.DataFrame:
    """
    Load historical market cap panel (date x ticker) from Parquet with standard normalization.

    Args:
      file_path: path to parquet file.
      tickers: optional list of tickers to keep (reduces memory).
      start/end: optional date filtering.
      sort_index: sort dates ascending.
      sort_columns: sort ticker columns.

    Returns:
      DataFrame indexed by DatetimeIndex, columns=tickers.
    """
    fp = Path(file_path)
    if not fp.exists():
        raise FileNotFoundError(fp)

    df = pd.read_parquet(fp)

    # Ensure DatetimeIndex
    if not isinstance(df.index, pd.DatetimeIndex):
        df.index = pd.to_datetime(df.index, errors="raise")

    # Optional date filtering
    if start is not None:
        df = df.loc[pd.Timestamp(start):]
    if end is not None:
        df = df.loc[:pd.Timestamp(end)]

    # Optional ticker filtering
    if tickers is not None:
        tickers = [str(t) for t in tickers]
        df = df.reindex(columns=tickers)

    if sort_index:
        df = df.sort_index()
    if sort_columns:
        df = df.reindex(sorted(df.columns), axis=1)

    return df


def save_estimated_shares(
    estimated_shares: pd.Series,
    file_path: str | Path,
    *,
    asof_date: str | pd.Timestamp | None = None,
    px_asof_date: str | pd.Timestamp | None = None,
    source: str = "yfinance_marketCap / latest_close",
    compression: str = "zstd",
) -> Path:
    """
    Save estimated shares outstanding as a small reference table.

    Args:
      estimated_shares: Series indexed by ticker, values = shares outstanding estimate.
      file_path: destination .parquet or .csv.
      asof_date: market cap snapshot date (metadata).
      px_asof_date: date of latest close used to compute shares (metadata).
      source: free-text provenance string.
      compression: parquet compression if saving parquet.

    Returns:
      Path to saved file.
    """
    fp = Path(file_path)
    fp.parent.mkdir(parents=True, exist_ok=True)

    s = estimated_shares.copy()
    s.index = s.index.astype(str)
    s.name = "estimated_shares"

    df = s.reset_index().rename(columns={"index": "ticker"})

    if asof_date is not None:
        df["asof_date"] = pd.Timestamp(asof_date).date().isoformat()
    if px_asof_date is not None:
        df["px_asof_date"] = pd.Timestamp(px_asof_date).date().isoformat()

    df["source"] = source

    if fp.suffix.lower() == ".csv":
        df.to_csv(fp, index=False)
    elif fp.suffix.lower() in {".parquet", ".pq"}:
        df.to_parquet(fp, index=False, compression=compression)
    else:
        raise ValueError("Unsupported file type. Use .csv or .parquet")

    return fp


def load_estimated_shares(file_path: str | Path) -> pd.Series:
    fp = Path(file_path)
    if fp.suffix.lower() == ".csv":
        df = pd.read_csv(fp)
    elif fp.suffix.lower() in {".parquet", ".pq"}:
        df = pd.read_parquet(fp)
    else:
        raise ValueError("Unsupported file type. Use .csv or .parquet")

    if 'ticker' in df.columns:
        ticker_col = 'ticker'
    elif 'Ticker' in df.columns:
        ticker_col = 'Ticker'
    else:
        raise KeyError("Expected columns: ticker, estimated_shares")

    if "estimated_shares" not in df.columns:
        raise KeyError("Expected columns: ticker, estimated_shares")

    s = df.set_index(ticker_col)["estimated_shares"].astype(float)
    s.name = "estimated_shares"

    return s


def finalize_mcap_snapshot(
        mkt_cp_df: pd.DataFrame,
        *,
        ticker_col: str = "Ticker",
        mcap_col: str = "Market Cap",
        asof_date: str | pd.Timestamp,
) -> pd.DataFrame:
    df = mkt_cp_df[[ticker_col, mcap_col]].copy()
    df.columns = ["ticker", "market_cap"]
    df["ticker"] = df["ticker"].astype(str)
    df["market_cap"] = pd.to_numeric(df["market_cap"], errors="coerce")
    df["asof_date"] = pd.Timestamp(asof_date).date().isoformat()

    # drop missing / non-positive
    df = df.dropna(subset=["market_cap"])
    df = df[df["market_cap"] > 0]

    # de-dup if yfinance returned duplicates
    df = df.drop_duplicates(subset=["ticker"], keep="last").reset_index(drop=True)
    return df


def save_mcap_snapshot(df: pd.DataFrame, path: str | Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(path, index=False, compression="zstd")

    return path


def compute_estimated_shares(
        close: pd.DataFrame,
        mktcap_snapshot: pd.DataFrame,
        *,
        px_asof: str | pd.Timestamp | None = None,
) -> pd.Series:
    if px_asof is None:
        px_asof = close.index[-1]
    px_asof = pd.Timestamp(px_asof)

    latest_px = close.loc[px_asof]
    snap = mktcap_snapshot.set_index("ticker")["market_cap"]

    shares = (snap / latest_px).rename("estimated_shares")
    shares = shares.replace([np.inf, -np.inf], np.nan)
    shares = shares[(shares > 0) & shares.notna()]
    return shares


def build_historical_mcap(
        close: pd.DataFrame,
        estimated_shares: pd.Series,
) -> pd.DataFrame:
    sh = estimated_shares.reindex(close.columns.astype(str))
    return close.mul(sh, axis=1)


def save_hsci_components_history(
        fname: str = 'history_hsci.xlsx',
        exclude: Optional[dict] = None,
        header_row_excel: int = 5
):
    """
        Main pipeline function to process the HSCI components history Excel file.

        It iterates through all non-excluded sheets in the raw Excel file, parses them
        using `read_constituents_sheet`, concatenates the results into a single DataFrame,
        and saves the final dataset to the processed directory as a CSV.

        Args:
            fname (str): Filename of the raw Excel file.
            exclude (Optional[dict]): A set or dict of sheet names to exclude from processing.
                                      Defaults to {'HSCI'} if None.
            header_row_excel (int): The 1-based row index in Excel where the header is located.
    """

    if exclude is None:
        exclude = {'HSCI'}
    _path = get_raw_dir() / fname

    print(f"--- Starting Build: {fname} ---")

    xls = pd.ExcelFile(_path, engine="openpyxl")

    sheets = [sh for sh in xls.sheet_names if sh not in exclude]

    all_df = pd.concat(
        [read_constituents_sheet(_path, sh, header_row_excel=header_row_excel).assign(sheet=sh) for sh in sheets],
        ignore_index=True,
    )
    print(f"Merge complete. Total shape: {all_df.shape}")
    print(f"Unique sheets successfully read: {all_df['sheet'].nunique()} / {len(sheets)}")

    all_df.to_csv(get_processed_dir() / 'hsci_components_history.csv')
    print(f"Saved to: {_path}")


def load_hsci_components_history(fname: str = 'hsci_components_history.csv'):
    """
        Utility function to load the processed HSCI components history data.

        Args:
            fname (str): The filename to look for in the processed directory.

        Returns:
            pd.DataFrame: The loaded components history data.
    """

    return pd.read_csv(get_processed_dir() / fname)


def load_hsci_2008_components(fname: str = "hsci_components_2008.csv") -> pd.DataFrame:
    """
    Load the 2008 HSCI snapshot from raw/.
    """

    return pd.read_csv(get_raw_dir() / fname)


def read_constituents_sheet(path: str, sheet_name: str, header_row_excel: int = 5) -> pd.DataFrame:
    header_idx = header_row_excel - 1
    df = pd.read_excel(path, sheet_name=sheet_name, header=header_idx, engine="openpyxl")

    df = df.dropna(how="all").dropna(axis=1, how="all")
    df.columns = [str(c).split("\n")[0].strip() for c in df.columns]

    group_cols = [c for c in GROUP_COL_CANDIDATES if c in df.columns]
    if group_cols:
        df[group_cols] = df[group_cols].ffill()

    if ID_COL in df.columns:
        df = df[df[ID_COL].notna()].copy()
        df[ID_COL] = (
            df[ID_COL].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
        )
        mask = df[ID_COL].str.fullmatch(r"\d+")
        df.loc[mask, ID_COL] = df.loc[mask, ID_COL].str.zfill(4)

    if "Count" in df.columns:
        df["Count"] = pd.to_numeric(
            df["Count"].astype(str).str.replace("+", "", regex=False),
            errors="coerce",
        )

    if "Change" in df.columns:
        df["Action"] = df["Change"].astype(str).str.extract(r"^(Add|Delete)", expand=False)

    # ---- sector from A2 ----
    sector = _extract_sector_from_a2(path, sheet_name)
    df["Sector"] = sector

    return df


def _extract_sector_from_a2(path: str, sheet_name: str) -> str | None:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    v = ws["A2"].value
    if v is None:
        return None
    v = str(v).strip()
    m = SECTOR_RE.search(v)
    return m.group(1).strip() if m else None


def long_to_matrix(
    df: pd.DataFrame,
    *,
    date_col: str,
    code_col: str,
    value_col: str,
    dates: pd.DatetimeIndex,
    tickers: pd.Index,
) -> pd.DataFrame:
    x = df[[date_col, code_col, value_col]].copy()
    x[date_col] = pd.to_datetime(x[date_col], errors="coerce").dt.normalize()
    x[code_col] = x[code_col].astype("string").str.strip()
    x = x.dropna(subset=[date_col, code_col])

    mat = x.pivot(index=date_col, columns=code_col, values=value_col)
    # Align to the canonical grid
    return mat.reindex(index=dates, columns=tickers)
